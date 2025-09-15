# ==============================================================================
#                 ZKTeco Attendance & SMS Notification System
# ==============================================================================
# FINAL DEPLOYMENT VERSION: Direct Device Punch Code Method
# This is the cleanest, fastest, and most accurate architecture.
# ==============================================================================

import openpyxl
from openpyxl import load_workbook, Workbook
import http.client
import json
import pyodbc
import ssl
from zk import ZK, const
import time
import mysql.connector
import datetime
import logging
import os
from logging.handlers import TimedRotatingFileHandler
import sys
import atexit
import threading
import shutil
import tempfile

# ==============================================================================
#                      --- CONFIGURATION SECTION ---
# ==============================================================================

# --- Device Configuration (MODIFY THIS FOR EACH SCRIPT COPY) ---
DEVICES = [
    # Example for a specific device script. This is the only line you change per copy.
    {'ip': '192.168.100.200', 'port': 4370, 'name': 'Device A'}, 
]

# --- Get a safe filename from the device name ---
DEVICE_NAME_SAFE = DEVICES[0]['name'].replace(' ', '_') if DEVICES else "default"

# --- Setup Logging ---
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)
log_formatter = logging.Formatter('%(asctime)s - %(process)d - %(levelname)s - %(message)s')
log_file = os.path.join(LOG_DIR, f"attendance_{DEVICE_NAME_SAFE}.log")
handler = TimedRotatingFileHandler(log_file, when="midnight", interval=1, backupCount=7)
handler.suffix = "%Y-%m-%d"
handler.setFormatter(log_formatter)
handler.setLevel(logging.INFO)
logger = logging.getLogger()
logger.setLevel(logging.INFO)
logger.addHandler(handler)
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
console_handler.setLevel(logging.INFO)
logger.addHandler(console_handler)

# --- Threading Lock for internal safety ---
excel_lock = threading.Lock()

# --- File Paths ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, f"attendance_log_{DEVICE_NAME_SAFE}.xlsx")
PID_FILE_PATH = os.path.join(SCRIPT_DIR, f"biometrics_{DEVICE_NAME_SAFE}.pid")

# --- Other Constants ---
SEMATIME_ACCOUNT_ID = "1374737440333"
SEMATIME_AUTH_TOKEN = "f5704c567d3d44199a1b61f1bf110719"
ACCESS_DB_PATH = r"C:\Program Files (x86)\ZKTeco\ZKTime5.0\att2000.mdb"
db_config = { 'host': 'localhost', 'user': 'root', 'password': '', 'database': 'attendance_systems' }
POLL_INTERVAL_SECONDS = 3
BATCH_SIZE = 10

# ==============================================================================
#                        --- CORE FUNCTIONS ---
# ==============================================================================

def create_pid_file():
    if os.path.isfile(PID_FILE_PATH):
        logger.error(f"PID file '{PID_FILE_PATH}' already exists. Aborting.")
        sys.exit(1)
    pid = os.getpid()
    try:
        with open(PID_FILE_PATH, 'w') as f: f.write(str(pid))
        logger.info(f"Script started with PID: {pid}. PID file created at {PID_FILE_PATH}")
    except IOError as e:
        logger.critical(f"Unable to create PID file '{PID_FILE_PATH}'. Error: {e}")
        sys.exit(1)

def remove_pid_file():
    if os.path.isfile(PID_FILE_PATH):
        try:
            os.remove(PID_FILE_PATH)
            logger.info(f"PID file removed on clean exit: {PID_FILE_PATH}")
        except IOError as e:
            logger.error(f"Unable to remove PID file '{PID_FILE_PATH}'. Error: {e}")

def _get_single_user_detail(user_id, conn):
    try:
        cursor = conn.cursor()
        query = "SELECT SSN, Name, OPHONE FROM USERINFO WHERE TRIM(SSN) = ?"
        cursor.execute(query, str(user_id))
        row = cursor.fetchone()
        if row:
            logger.info(f"SUCCESS on fallback: Found details for user {user_id}.")
            return {'name': str(row.Name).strip() if row.Name else None, 'phone': str(row.OPHONE).strip() if row.OPHONE else None}
        logger.warning(f"FAIL on fallback: User {user_id} still not found even with a targeted query.")
        return None
    except Exception as e:
        logger.error(f"Error during single user fallback for {user_id}: {e}")
        return None

def get_user_details_in_batch(user_ids):
    if not user_ids: return {}
    conn = None
    temp_dir_path = None
    user_details = {}
    try:
        temp_dir_path = tempfile.mkdtemp(prefix="access_db_batch_")
        temp_db_path = os.path.join(temp_dir_path, "att2000.mdb")
        shutil.copy2(ACCESS_DB_PATH, temp_db_path)
        temp_conn_str = (f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={temp_db_path};READONLY=TRUE;")
        conn = pyodbc.connect(temp_conn_str)
        safe_user_ids = [uid for uid in user_ids if uid.isdigit()]
        if not safe_user_ids: return {}
        placeholders = ', '.join(['?'] * len(safe_user_ids))
        query = f"SELECT SSN, Name, OPHONE FROM USERINFO WHERE SSN IN ({placeholders})"
        cursor = conn.cursor()
        cursor.execute(query, safe_user_ids)
        for row in cursor.fetchall():
            user_id_str = str(row.SSN)
            user_details[user_id_str] = {'name': str(row.Name).strip() if row.Name else None, 'phone': str(row.OPHONE).strip() if row.OPHONE else None}
        logger.info(f"Initial batch query found {len(user_details)} of {len(user_ids)} requested users.")
        found_ids = set(user_details.keys())
        missing_ids = set(user_ids) - found_ids
        if missing_ids:
            logger.warning(f"Initial batch failed for {len(missing_ids)} users. Trying one-by-one fallback for: {missing_ids}")
            for missing_id in missing_ids:
                fallback_detail = _get_single_user_detail(missing_id, conn)
                if fallback_detail:
                    user_details[missing_id] = fallback_detail
        return user_details
    except Exception as e:
        logger.error(f"General error in get_user_details_in_batch: {e}", exc_info=True)
        return {}
    finally:
        if conn: conn.close()
        if temp_dir_path and os.path.exists(temp_dir_path):
            try: shutil.rmtree(temp_dir_path)
            except OSError as e: logger.warning(f"Failed to delete temp Access DB batch directory {temp_dir_path}: {e}")

def connect_to_db():
    try: return mysql.connector.connect(**db_config)
    except mysql.connector.Error as err:
        logger.error(f"MySQL Connection Error: {err}")
        return None

def send_sms(message, recipient):
    if not recipient or not recipient.strip():
        logger.error("SMS Error: Recipient phone number is missing or empty.")
        return
    conn = http.client.HTTPSConnection("apis.sematime.com")
    payload = json.dumps({"message": message, "recipients": recipient})
    headers = {'Content-Type': 'application/json', 'AuthToken': SEMATIME_AUTH_TOKEN}
    try:
        url = f"/v1/{SEMATIME_ACCOUNT_ID}/messages/single"
        conn.request("POST", url, body=payload, headers=headers)
        res = conn.getresponse()
        data = res.read()
        logger.info(f"SMS API Response: {data.decode('utf-8')}")
        store_sms(message)
    except Exception as e: logger.error(f"SMS Sending Error: {e}")
    finally: conn.close()

def store_sms(message):
    connection = connect_to_db()
    if not connection: return
    try:
        cursor = connection.cursor()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        query = "INSERT INTO sms_logs (message, timestamp) VALUES (%s, %s)"
        cursor.execute(query, (message, timestamp))
        connection.commit()
    except mysql.connector.Error as err: logger.error(f"DB SMS Log Error: {err}")
    finally: connection.close()

def get_last_timestamp_from_db(device_name):
    connection = connect_to_db()
    if not connection: return None
    try:
        cursor = connection.cursor()
        query = "SELECT MAX(timestamp) FROM attendance_logs WHERE device_name = %s"
        cursor.execute(query, (device_name,))
        result = cursor.fetchone()
        if result and result[0]:
            logger.info(f"Last processed timestamp for '{device_name}' is {result[0]}.")
            return result[0]
        else:
            logger.info(f"No previous logs for '{device_name}'. Will process all.")
            return datetime.datetime(2000, 1, 1)
    except mysql.connector.Error as err:
        logger.error(f"DB Error getting last timestamp for '{device_name}': {err}")
        return None
    finally:
        if connection and connection.is_connected(): connection.close()

def process_log_batch(log_batch, device_info):
    device_name = device_info['name']
    logger.info(f"Processing batch of {len(log_batch)} logs for '{device_name}'...")
    unique_user_ids = list({str(log.user_id) for log in log_batch})
    user_details_cache = get_user_details_in_batch(unique_user_ids)
    excel_rows, db_records, sms_to_send = [], [], []

    for log in log_batch:
        logger.info(f"[{device_name}] RAW LOG DATA -> UserID: {log.user_id}, Timestamp: {log.timestamp}, PUNCH: {log.punch}")
        user_id = str(log.user_id)
        details = user_details_cache.get(user_id)
        if not details: continue
        
        user_name = details.get('name') or f"Name-Not-Found ({user_id})"
        phone_number = details.get('phone')
        
        ACTION_RETURN = "has reported back to school"
        ACTION_DEPART = "has left the school"
        
        # Logic based on ZKTeco punch codes: 0 = Check-In, 1 = Check-Out
        if log.punch == 0:
            action_str = ACTION_RETURN
        elif log.punch == 1:
            action_str = ACTION_DEPART
        else:
            logger.warning(f"Skipping log for user {user_id} due to unknown punch type: {log.punch}")
            continue
            
        display_timestamp_str = log.timestamp.strftime("%d-%m-%Y %I:%M %p")
        db_timestamp_str = log.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        excel_rows.append([device_name, user_id, user_id, user_name, action_str, display_timestamp_str])
        db_records.append((user_id, user_name, action_str, db_timestamp_str, device_name))

        if phone_number:
            message = f"Dear Parent/Guardian, your daughter, {user_name} ADM NO: {user_id} , {action_str} on {display_timestamp_str}."
            sms_to_send.append({'message': message, 'recipient': phone_number})

    with excel_lock:
        try:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook(); sheet = workbook.active
            sheet.append(["Device", "User ID", "Admission No", "Name", "Action", "Timestamp"])
        for row_data in excel_rows: sheet.append(row_data)
        workbook.save(EXCEL_FILE)
        logger.info(f"Appended {len(excel_rows)} rows to local Excel file: {EXCEL_FILE}")

    connection = connect_to_db()
    if connection and db_records:
        try:
            cursor = connection.cursor()
            query = "INSERT INTO attendance_logs (user_id, user_name, action, timestamp, device_name) VALUES (%s, %s, %s, %s, %s)"
            cursor.executemany(query, db_records)
            connection.commit()
            logger.info(f"Batch inserted {cursor.rowcount} records into MySQL for '{device_name}'.")
        except mysql.connector.Error as err: logger.error(f"MySQL batch insert error for '{device_name}': {err}")
        finally: connection.close()

    if sms_to_send:
        logger.info(f"Sending {len(sms_to_send)} SMS messages for '{device_name}'...")
        for sms_data in sms_to_send:
            send_sms(sms_data['message'], sms_data['recipient'])
            time.sleep(0.2)

def device_polling_thread(device_info):
    name = device_info['name']
    last_processed_timestamp = get_last_timestamp_from_db(name)
    if last_processed_timestamp is None:
        logger.critical(f"Could not retrieve last timestamp for '{name}'. Process exiting.")
        return
    while True:
        conn = None
        try:
            logger.info(f"Polling '{name}' ({device_info['ip']}) for new logs...")
            zk = ZK(device_info['ip'], port=device_info['port'], timeout=20, force_udp=False)
            conn = zk.connect()
            conn.disable_device()
            
            attendance_logs = conn.get_attendance()
            new_logs = [log for log in attendance_logs if log.timestamp > last_processed_timestamp]

            if new_logs:
                new_logs.sort(key=lambda x: x.timestamp)
                total_new = len(new_logs)
                logger.info(f"Found {total_new} new log(s) on '{name}'. Processing...")
                original_log_list = list(new_logs)
                while new_logs:
                    batch_to_process = new_logs[:BATCH_SIZE]
                    new_logs = new_logs[BATCH_SIZE:]
                    process_log_batch(batch_to_process, device_info)
                    logger.info(f"Batch for '{name}' finished. Remaining logs to process: {len(new_logs)}")
                
                last_processed_timestamp = original_log_list[-1].timestamp
                logger.info(f"Finished ALL batches for '{name}'. New last timestamp is {last_processed_timestamp}.")
            else:
                logger.info(f"No new logs on '{name}'.")
        except Exception as e:
            logger.error(f"An error occurred in the polling loop for '{name}': {e}", exc_info=True)
        finally:
            if conn and conn.is_connect:
                try:
                    conn.enable_device()
                    logger.info(f"Successfully re-enabled device '{name}'.")
                except Exception as e: logger.error(f"Could not re-enable device '{name}' during cleanup: {e}")
                try:
                    conn.disconnect()
                    logger.info(f"Successfully disconnected from '{name}'.")
                except Exception as e: logger.error(f"Error during disconnection from '{name}': {e}")
            
            logger.info(f"Polling cycle for '{name}' complete. Waiting {POLL_INTERVAL_SECONDS} seconds.")
            time.sleep(POLL_INTERVAL_SECONDS)

def main():
    if not DEVICES:
        logger.error("No devices configured in the DEVICES list. Exiting.")
        return
    
    device_info = DEVICES[0]
    logger.info(f"Starting dedicated polling process for '{device_info['name']}'.")
    
    device_polling_thread(device_info)

if __name__ == "__main__":
    atexit.register(remove_pid_file)
    create_pid_file()
    try:
        main()
    except KeyboardInterrupt:
        logger.info("Shutdown signal received. Exiting gracefully.")
    except Exception as e:
        logger.critical(f"A critical unhandled error occurred in main execution: {e}", exc_info=True)
        sys.exit(1)